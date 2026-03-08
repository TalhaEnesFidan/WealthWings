"""
PDF ve Excel dışa aktarım işlemleri
"""
import io
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import extract
from datetime import datetime

from backend.database import get_db
from backend.models import Transaction, Category, User
from backend.routers.auth import get_current_user

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm

router = APIRouter()


@router.get("/excel")
def export_excel(
    year: int = Query(None),
    month: int = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Kullanıcının işlemlerini (filtreli veya tümü) Excel olarak dışa aktarır"""
    
    query = db.query(Transaction).filter(Transaction.user_id == current_user.id)
    
    if year:
        query = query.filter(extract('year', Transaction.transaction_date) == year)
    if month:
        query = query.filter(extract('month', Transaction.transaction_date) == month)
        
    transactions = query.order_by(Transaction.transaction_date.desc()).all()
    
    # Excel dosyası oluştur
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "İşlemler"
    
    # Başlıklar
    headers = ["Tarih", "Tip", "Kategori", "Tutar (₺)", "Açıklama"]
    ws.append(headers)
    
    # Stil
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    for col_num in range(1, 6):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    # Veriler
    for t in transactions:
        cat_name = "Kategorisiz"
        if t.category_id:
            cat = db.query(Category).filter(Category.id == t.category_id).first()
            if cat:
                cat_name = cat.name
                
        tip_str = "Gelir" if t.type == "income" else "Gider"
        
        ws.append([
            t.transaction_date.strftime("%d.%m.%Y"),
            tip_str,
            cat_name,
            float(t.amount),
            t.description or ""
        ])
        
    # Sütun genişlikleri
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 40
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"wealthwings_islemler_{datetime.now().strftime('%Y%m%d')}.xlsx"
    if year and month:
        filename = f"wealthwings_{year}_{month:02d}_islemler.xlsx"
        
    headers_res = {
        'Content-Disposition': f'attachment; filename="{filename}"'
    }
    
    return StreamingResponse(
        output, 
        headers=headers_res, 
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@router.get("/pdf")
def export_pdf(
    year: int = Query(None),
    month: int = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Kullanıcının aylık veya yıllık özetini PDF olarak dışa aktarır"""
    if not year:
        year = datetime.now().year
    if not month:
        month = datetime.now().month
        
    query = db.query(Transaction).filter(
        Transaction.user_id == current_user.id,
        extract('year', Transaction.transaction_date) == year,
        extract('month', Transaction.transaction_date) == month
    )
    
    transactions = query.order_by(Transaction.transaction_date.desc()).all()
    
    # PDF oluştur
    output = io.BytesIO()
    c = canvas.Canvas(output, pagesize=A4)
    width, height = A4
    
    # Başlık
    c.setFont("Helvetica-Bold", 20)
    c.drawString(2 * cm, height - 2 * cm, "WealthWings Finansal Özet Raporu")
    
    c.setFont("Helvetica", 12)
    c.drawString(2 * cm, height - 3 * cm, f"Kullanıcı: {current_user.full_name or current_user.username}")
    c.drawString(2 * cm, height - 3.5 * cm, f"Dönem: {month:02d}/{year}")
    
    # Toplamları hesapla
    total_inc = sum(t.amount for t in transactions if t.type == "income")
    total_exp = sum(t.amount for t in transactions if t.type == "expense")
    net = total_inc - total_exp
    
    c.setFont("Helvetica-Bold", 14)
    c.drawString(2 * cm, height - 5 * cm, f"Toplam Gelir: {total_inc:,.2f} TL")
    c.drawString(2 * cm, height - 5.5 * cm, f"Toplam Gider: {total_exp:,.2f} TL")
    
    if net >= 0:
        c.setFillColorRGB(0.13, 0.77, 0.36)  # success green #22c55e
    else:
        c.setFillColorRGB(0.93, 0.26, 0.26)  # danger red #ef4444
        
    c.drawString(2 * cm, height - 6 * cm, f"Net Nakit Akışı: {net:,.2f} TL")
    c.setFillColorRGB(0, 0, 0)
    
    # İşlemler Listesi
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2 * cm, height - 8 * cm, "Son İşlemler:")
    
    y = height - 9 * cm
    c.setFont("Helvetica", 10)
    for t in transactions[:25]: # Sayfaya sığması için ilk 25 işlem (basit PDF)
        tip = "Gelir" if t.type == "income" else "Gider"
        c.drawString(2 * cm, y, f"{t.transaction_date.strftime('%d.%m.%Y')} | {tip} | {t.amount:,.2f} TL | {t.description[:30]}")
        y -= 0.6 * cm
        
        if y < 2 * cm:
            c.showPage()
            y = height - 2 * cm
            c.setFont("Helvetica", 10)

    # Rapor oluşturma tarihi
    c.setFont("Helvetica-Oblique", 8)
    c.drawString(2 * cm, 1 * cm, f"Oluşturulma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    
    c.save()
    output.seek(0)
    
    filename = f"wealthwings_{year}_{month:02d}_ozet.pdf"
    headers_res = {
        'Content-Disposition': f'attachment; filename="{filename}"'
    }
    
    return StreamingResponse(
        output,
        headers=headers_res,
        media_type='application/pdf'
    )
